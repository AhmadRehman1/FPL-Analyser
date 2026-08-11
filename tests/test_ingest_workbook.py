from pathlib import Path

import openpyxl
import pytest

from fpl_quant import ingest_workbook as iw

REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = REPO_ROOT / "data" / "external" / "FPL_202627_Master_Evidence_Database.xlsx"

pytestmark = pytest.mark.skipif(not XLSX_PATH.exists(), reason="real evidence workbook not available")


@pytest.fixture(scope="module")
def wb():
    return openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)


def test_every_tab_is_classified(wb):
    unclassified = set(wb.sheetnames) - iw.ALL_CLASSIFIED
    assert unclassified == set()


def test_original_m0_deprecation_list_present():
    expected = {
        "1_Player Database", "2_Team Database", "3_Fixture Database",
        "21_Master Player Rating Engine", "21b_Rating Engine Methodology",
        "28_Master Rating Engine v2", "28b_Rating Engine v2 Method",
        "29_Position Rankings", "30_Fixture Swing GW1-10",
    }
    assert expected.issubset(iw.EXCLUDED_DEPRECATED)


def test_m8_tab_31_36_addition_present():
    # kickoff notes item 4: tabs 31-36 are transitively built on the deprecated
    # 28_Master Rating Engine v2 and must be added to M0's exclusion allowlist.
    expected = {
        "31_Captaincy-Transfer Plan", "32_Chip Strategy Plan", "33_10 Squad Variants",
        "34_Sensitivity Analysis", "35_Risk Report", "36_Top-N Final Lists",
    }
    assert expected.issubset(iw.EXCLUDED_DEPRECATED)


def test_deprecated_and_ingestible_sets_are_disjoint():
    assert iw.EXCLUDED_DEPRECATED.keys() & iw.INGEST_CLAIMS == set()
    assert iw.EXCLUDED_DEPRECATED.keys() & iw.INGEST_MANUAL_DECOMPOSITION == set()


def test_seed_allowlist_populates_every_tab(con, wb):
    iw.seed_allowlist(con, wb)
    n = con.execute("SELECT count(*) FROM workbook_tab_allowlist").fetchone()[0]
    assert n == len(wb.sheetnames)


def test_seed_allowlist_marks_deprecated_tabs_correctly(con, wb):
    iw.seed_allowlist(con, wb)
    for tab in iw.EXCLUDED_DEPRECATED:
        status = con.execute(
            "SELECT status FROM workbook_tab_allowlist WHERE tab_name = ?", [tab]
        ).fetchone()[0]
        assert status == "excluded_deprecated"


def test_deprecated_tabs_never_produce_evidence_claims(con, wb):
    # structural guarantee: ingest_all() has no code path that reads a deprecated tab for
    # claims -- confirmed by checking every evidence_claims.tab_origin actually written.
    from fpl_quant import params as params_mod

    params_mod.write_param(
        con, "source_tier_weights", 1, "2026-08-10", "tier_weight",
        value_numeric=0.4, dimensions={"source_type": "community"},
    )
    iw.ingest_all(con, str(XLSX_PATH), source_tier_params_version=1)
    origins = {r[0] for r in con.execute("SELECT DISTINCT tab_origin FROM evidence_claims").fetchall()}
    assert origins & set(iw.EXCLUDED_DEPRECATED) == set()


def test_classify_source_type_known_official():
    assert iw.classify_source_type("PL") == "official"
    assert iw.classify_source_type("Premier League") == "official"


def test_classify_source_type_journalist_keyword():
    assert iw.classify_source_type("BBC/Yahoo Sports/Guardian") == "journalist"


def test_classify_source_type_specialist_keyword():
    assert iw.classify_source_type("Fantasy Football Fix") == "specialist"
    assert iw.classify_source_type("RRI") == "specialist"


def test_classify_source_type_unknown_defaults_to_community():
    assert iw.classify_source_type("Some Random Twitter Account") == "community"


def test_confidence_normalizes_1_to_10_scale():
    assert iw._confidence_0_1(9) == 0.9
    assert iw._confidence_0_1(None) is None
    assert iw._confidence_0_1("") is None
