"""Turn a Perplexity evidence-pull answer (markdown tables, docs/evidence_research_pull_prompt.md
format) into the FPL_Evidence_Claims_Research_Pull.xlsx workbook that scripts/run_ingestion.py
feeds to ingest_research_pull.py.

  PYTHONPATH=src python scripts/build_evidence_pull_workbook.py <perplexity_answer.md> [out.xlsx]

What it does:
  * finds every "## Table N -- <Name>" section and parses the markdown table under it
  * maps the six table names to the sheet names ingest_research_pull expects
  * expands bare surnames in player / taker columns to canonical full names, using FPL's own
    bootstrap-static (web_name + "first second"), so the ingest's normalized-name resolver can
    match them -- ambiguous surnames (two+ players) are left as-is and listed in the report
  * normalises the injury `status` column to the categories minutes_adjustment_params keys off
    (Out / Doubt / Doubt (improving) / Fit / Minor/knock), keeping the raw text in `notes`
  * writes the .xlsx and prints a validation report (rows per sheet, unresolved names)

Network: fetches fantasy.premierleague.com/api/bootstrap-static/ for the name map. If that's
blocked, pass --no-resolve to skip surname expansion (the ingest will then drop bare-surname
rows, which it reports as skipped).
"""

from __future__ import annotations

import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]

# markdown "## Table N -- Name"  ->  workbook sheet name ingest_research_pull.py reads
_TABLE_TO_SHEET = {
    "injuries": "Injuries",
    "predictedxi": "PredictedXI",
    "predicted xi": "PredictedXI",
    "rotation": "Rotation",
    "rolechange": "RoleChange",
    "role change": "RoleChange",
    "setpieces": "SetPieces",
    "set pieces": "SetPieces",
    "pricewatch": "PriceWatch",
    "price watch": "PriceWatch",
}

# player-name columns per sheet that need surname -> full-name expansion
_NAME_COLS = {
    "Injuries": ["player"],
    "PredictedXI": ["player"],
    "Rotation": ["player"],
    "RoleChange": ["player"],
    "SetPieces": ["primary_taker", "secondary_taker", "deputy_if_primary_absent"],
    "PriceWatch": ["player"],
}


def _parse_markdown_tables(text: str) -> dict[str, list[dict]]:
    """{sheet_name: [row_dict, ...]} for every recognised '## Table ... -- Name' section."""
    out: dict[str, list[dict]] = {}
    # split on headings that look like a table section
    sections = re.split(r"\n#{2,3}\s+", "\n" + text)
    for sec in sections:
        head, _, body = sec.partition("\n")
        m = re.search(r"table\s*\d*\s*[-—:]+\s*(.+)", head, re.IGNORECASE)
        if not m:
            # also accept a bare "Name" heading immediately above a table
            name_key = head.strip().lower()
        else:
            name_key = re.sub(r"\(.*?\)", "", m.group(1)).strip().lower()
        sheet = _TABLE_TO_SHEET.get(name_key)
        if not sheet:
            continue
        rows = _parse_one_table(body)
        if rows:
            out.setdefault(sheet, []).extend(rows)
    return out


def _parse_one_table(body: str) -> list[dict]:
    lines = [ln.rstrip() for ln in body.splitlines()]
    tbl = [ln for ln in lines if ln.strip().startswith("|")]
    if len(tbl) < 2:
        return []

    def cells(ln: str) -> list[str]:
        parts = [c.strip() for c in ln.strip().strip("|").split("|")]
        return parts

    header = [re.sub(r"[\s\-]+", "_", h.lower()) for h in cells(tbl[0])]
    rows = []
    for ln in tbl[1:]:
        if re.fullmatch(r"[\s|:\-]+", ln):  # the |---|---| separator
            continue
        vals = cells(ln)
        if len(vals) != len(header):
            continue
        row = {header[i]: _clean_cell(vals[i]) for i in range(len(header))}
        if any(v for v in row.values()):
            rows.append(row)
    return rows


def _clean_cell(v: str) -> str:
    v = re.sub(r"\[\^\d+\]", "", v)          # drop [^1] footnote markers
    v = v.replace("**", "").replace("`", "").strip()
    return "" if v in ("-", "—", "–", "n/a", "N/A") else v


def _fetch_fpl_name_map() -> tuple[dict[str, str], set[str]]:
    """(name_variant_lower -> 'First Last', ambiguous_surnames). Best effort."""
    req = urllib.request.Request(
        "https://fantasy.premierleague.com/api/bootstrap-static/", headers={"User-Agent": "Mozilla/5.0"}
    )
    data = json.load(urllib.request.urlopen(req, timeout=25))
    full_by_surname: dict[str, set[str]] = {}
    variants: dict[str, str] = {}
    for e in data["elements"]:
        first, second, web = e.get("first_name", ""), e.get("second_name", ""), e.get("web_name", "")
        full = f"{first} {second}".strip()
        if not full:
            continue
        variants.setdefault(full.lower(), full)
        if web:
            variants.setdefault(web.lower(), full)
        surname = second.split()[-1].lower() if second else ""
        if surname:
            full_by_surname.setdefault(surname, set()).add(full)
    ambiguous = {s for s, fs in full_by_surname.items() if len(fs) > 1}
    for s, fs in full_by_surname.items():
        if len(fs) == 1:
            variants.setdefault(s, next(iter(fs)))
    return variants, ambiguous


_STATUS_MAP = [
    (r"^out\b|suspend|red card|ban\b|acl|season-ending", "Out"),
    (r"improv|edging|close to a return|managed return", "Doubt (improving)"),
    (r"minor|knock", "Minor/knock"),
    (r"doubt|75%|50%|25%|questionable", "Doubt"),
    (r"^fit\b|returned|available", "Fit"),
]


def _normalise_status(raw: str) -> str:
    low = raw.lower()
    for pat, cat in _STATUS_MAP:
        if re.search(pat, low):
            return cat
    return raw or "Doubt"


def build(md_path: Path, out_path: Path, *, resolve: bool = True) -> dict:
    text = md_path.read_text(encoding="utf-8")
    sheets = _parse_markdown_tables(text)
    if not sheets:
        raise SystemExit(f"no recognised markdown tables found in {md_path}")

    variants: dict[str, str] = {}
    ambiguous: set[str] = set()
    if resolve:
        try:
            variants, ambiguous = _fetch_fpl_name_map()
        except Exception as exc:  # noqa: BLE001
            print(f"::warning::name resolution skipped ({exc}); bare surnames will not be expanded")

    unresolved: list[str] = []

    def canon(name: str) -> str:
        if not name:
            return name
        key = name.lower()
        if key in variants:
            return variants[key]
        # try last token as a surname
        toks = re.sub(r"[^\w\s]", " ", key).split()
        if toks and toks[-1] in variants and toks[-1] not in ambiguous:
            return variants[toks[-1]]
        if len(name.split()) == 1 and resolve:
            unresolved.append(name)
        return name

    for sheet, rows in sheets.items():
        for row in rows:
            for col in _NAME_COLS.get(sheet, []):
                if col in row and row[col]:
                    row[col] = canon(row[col])
            if sheet == "Injuries" and "status" in row:
                raw = row["status"]
                row["status"] = _normalise_status(raw)
                if row["status"] != raw:
                    row["notes"] = (row.get("notes", "") + f" [raw status: {raw}]").strip()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    readme.append(["FPL Quant -- Evidence Claims Database (built by scripts/build_evidence_pull_workbook.py)"])
    readme.append([f"source: {md_path.name}"])
    for sheet, rows in sheets.items():
        ws = wb.create_sheet(sheet)
        headers = list({k for row in rows for k in row})
        # stable, human order: put the well-known columns first
        preferred = ["player", "club", "status", "issue", "predicted_starter", "start_confidence_pct",
                     "expected_minutes", "position", "valence", "pattern", "trigger", "manager",
                     "change", "cause", "effective_from", "duty", "primary_taker", "secondary_taker",
                     "deputy_if_primary_absent", "direction", "note", "date_reported", "expected_return",
                     "source_name", "source_type", "confidence_1_10", "information_type", "observed_date", "notes"]
        headers = [h for h in preferred if h in headers] + [h for h in headers if h not in preferred]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    report = {
        "out": str(out_path),
        "rows_per_sheet": {s: len(r) for s, r in sheets.items()},
        "unresolved_single_names": sorted(set(unresolved)),
    }
    return report


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    md_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else (
        REPO_ROOT / "data" / "external" / "FPL_Evidence_Claims_Research_Pull.xlsx"
    )
    resolve = "--no-resolve" not in sys.argv
    report = build(md_path, out_path, resolve=resolve)
    print(json.dumps(report, indent=2))
    if report["unresolved_single_names"]:
        print("\nUnresolved single-word names (left as-is; the ingest will drop these rows unless "
              "they happen to resolve): " + ", ".join(report["unresolved_single_names"]))


if __name__ == "__main__":
    main()
