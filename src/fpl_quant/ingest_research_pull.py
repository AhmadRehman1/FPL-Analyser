"""Ingests the manual web-research evidence pull into evidence_claims -- a second,
differently-shaped evidence source alongside the 40-tab master workbook (flat columns per
tab, one fact type per tab). Same destination schema (evidence_claims), same source-tier
classification and M1b reliability scoring.

Two generations of workbook shape are read, whichever tabs are present:

  v2 (docs/evidence_research_pull_prompt.md -- the Perplexity weekly-refresh format):
    Injuries     -> injury_status        (status Out/Doubt/Fit drives the minutes-model shift)
    PredictedXI  -> predicted_xi         (start_confidence_pct -> the minutes-model start pull)
    Rotation     -> manager_tendency     (valence positive/negative -> +/- minutes shift)
    RoleChange   -> manager_tendency / transfer_likelihood, routed by the `change` value
    SetPieces    -> set_piece_order_override (primary penalty/FK/corner duty -> EP uplift)
    PriceWatch   -> fpl_price_note        (audit/provenance only, never a real now_cost)

  v1 (the original 4-tab pull, still accepted so an older workbook keeps ingesting):
    Transfers -> transfer_likelihood · Injuries(v1) -> injury_status
    SetPieceTakers -> set_piece_order_override · PriceNotes -> fpl_price_note

Every column is read by header name, not position, so an extra column or a reordering in the
sheet never shifts the parse. A malformed row is counted as skipped, never raised -- this
runs inside scripts/run_ingestion.py, and one bad research row must not take down the whole
M0-M6 pipeline (and with it every downstream workflow).

PriceWatch / PriceNotes are ingested for provenance visibility only (claim_type=
'fpl_price_note') -- never promoted into fact_reconciled's authoritative now_cost. That
boundary is M0's own principle: a stat is never conflated with an opinion, however well
sourced.
"""

import json
import re
from datetime import date, datetime, timezone

import duckdb
import openpyxl

from . import ingest_workbook as iw

# Every tab name this module knows how to read, across both workbook generations. Anything
# else in the workbook (a README, a changelog) is ignored without comment.
_KNOWN_TABS = (
    "Injuries", "PredictedXI", "Rotation", "RoleChange", "SetPieces", "PriceWatch",  # v2
    "Transfers", "SetPieceTakers", "PriceNotes",                                     # v1 only
)


def _parse_flexible_date(val) -> date | None:
    """Research-note dates are messy: '2026-08-10' (clean), '2026-08-xx' (unknown day),
    '2026-summer' (unknown month). Approximated to a representative day -- mid-month for a
    known month, mid-July for a bare season/summer reference -- rather than dropped."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{4})-(\d{1,2})-xx$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 15)
    m = re.match(r"^(\d{4})-(summer|winter)$", s, re.IGNORECASE)
    if m:
        return date(int(m.group(1)), 7 if m.group(2).lower() == "summer" else 1, 15)
    return None


def _rows(wb: openpyxl.Workbook, tab: str) -> list[dict]:
    """Sheet -> list of {normalized_header: value}. Normalized = lowercased, spaces/hyphens
    collapsed to '_'. Missing tab -> []."""
    if tab not in wb.sheetnames:
        return []
    ws = wb[tab]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return []
    keys = [re.sub(r"[\s\-]+", "_", str(h).strip().lower()) if h is not None else f"_col{i}"
            for i, h in enumerate(header)]
    out = []
    for raw in it:
        row = {keys[i]: (raw[i] if i < len(raw) else None) for i in range(len(keys))}
        if any(v not in (None, "") for v in row.values()):
            out.append(row)
    return out


def _s(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float | None:
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def register_new_sources(con: duckdb.DuckDBPyConnection, wb: openpyxl.Workbook, params_version: int) -> int:
    existing = {r[0] for r in con.execute("SELECT source_name FROM sources").fetchall()}
    max_citations = con.execute("SELECT max(citation_count) FROM sources").fetchone()[0] or 1

    names: set[str] = set()
    for tab in _KNOWN_TABS:
        for row in _rows(wb, tab):
            name = _s(row.get("source_name") or row.get("source"))
            if name:
                names.add(name)

    n = 0
    for name in sorted(names - existing):
        source_type = iw.classify_source_type(name)
        row = con.execute(
            "SELECT value_numeric FROM param_versions WHERE param_family = 'source_tier_weights' "
            "AND param_version = ? AND param_key = 'tier_weight' AND dimensions = ?",
            [params_version, json.dumps({"source_type": source_type}, sort_keys=True, separators=(",", ":"))],
        ).fetchone()
        tier_weight = row[0] if row else 0.4
        base_score = tier_weight * iw._log_scaled(1, max_citations)
        source_id = "src_" + name.lower().replace(" ", "_").replace("/", "_")
        con.execute(
            "INSERT INTO sources (source_id, source_name, source_type, base_reliability_score, "
            "citation_count, source_notes, last_reviewed_date) VALUES (?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT (source_name) DO NOTHING",
            [source_id, name, source_type, base_score, 1, None, None],
        )
        n += 1
    return n


def _common(con, row: dict) -> dict | None:
    """The source/date/confidence/information_type fields every row carries, or None if the
    row can't be attributed to a registered source."""
    source_id = iw._source_id_for(con, _s(row.get("source_name") or row.get("source")))
    if not source_id:
        return None
    info = (_s(row.get("information_type")) or "OPINION").upper()
    return {
        "source_id": source_id,
        "source_reliability_score": iw._reliability_for(con, source_id),
        "confidence": iw._confidence_0_1(row.get("confidence_1_10") or row.get("confidence")),
        "information_type": "FACT" if info.startswith("FACT") else "OPINION",
        "observed_date": _parse_flexible_date(row.get("observed_date") or row.get("source_date") or row.get("date")),
    }


def _emit(con, ingested_date, *, player_uid, claim_type, claim_value, common, tab, i, numeric=None) -> bool:
    return iw._insert_claim(
        con, subject_entity_type="player", subject_entity_id=player_uid, claim_type=claim_type,
        claim_value=claim_value, claim_value_numeric=numeric,
        information_type=common["information_type"], source_id=common["source_id"],
        source_reliability_score=common["source_reliability_score"], confidence=common["confidence"],
        observed_date=common["observed_date"], ingested_date=ingested_date,
        tab_origin=f"research_pull:{tab}", row_origin=i,
    )


# ------------------------------------------------------------------ per-tab handlers

def ingest_injuries(con, wb, ingested_date, tab="Injuries") -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, tab), start=2):
        name = _s(row.get("player") or row.get("subject") or row.get("subject_entity"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        if not uid or not common:
            skip += 1
            continue
        status = _s(row.get("status")) or "Doubt"
        payload = {"category": status, "issue": _s(row.get("issue") or row.get("claim_value")),
                   "date_reported": _s(row.get("date_reported")), "expected_return": _s(row.get("expected_return")),
                   "notes": _s(row.get("notes"))}
        ins += _emit(con, ingested_date, player_uid=uid, claim_type="injury_status",
                     claim_value=payload, common=common, tab=tab, i=i)
    return {"inserted": ins, "skipped": skip}


def ingest_predicted_xi(con, wb, ingested_date) -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, "PredictedXI"), start=2):
        name = _s(row.get("player") or row.get("subject"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        if not uid or not common:
            skip += 1
            continue
        pct = _num(row.get("start_confidence_pct") or row.get("starting_confidence_pct"))
        numeric = None if pct is None else max(0.0, min(1.0, pct / 100.0))
        payload = {"predicted_starter": _s(row.get("predicted_starter")),
                   "expected_minutes": _s(row.get("expected_minutes")),
                   "position": _s(row.get("position")), "notes": _s(row.get("notes"))}
        ins += _emit(con, ingested_date, player_uid=uid, claim_type="predicted_xi",
                     claim_value=payload, common=common, tab="PredictedXI", i=i, numeric=numeric)
    return {"inserted": ins, "skipped": skip}


def ingest_rotation(con, wb, ingested_date) -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, "Rotation"), start=2):
        name = _s(row.get("player") or row.get("subject"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        valence = (_s(row.get("valence")) or "").lower()
        if not uid or not common or valence not in ("positive", "negative"):
            skip += 1  # a rotation claim with no valence has no sign for minutes_model -- skip, never default
            continue
        payload = {"valence": valence, "manager": _s(row.get("manager")),
                   "pattern": _s(row.get("pattern")), "trigger": _s(row.get("trigger")),
                   "notes": _s(row.get("notes"))}
        ins += _emit(con, ingested_date, player_uid=uid, claim_type="manager_tendency",
                     claim_value=payload, common=common, tab="Rotation", i=i)
    return {"inserted": ins, "skipped": skip}


# change value -> (claim_type, manager_tendency valence). new_position is logged with no
# minutes effect (manager_tendency with an unrecognised valence -> sign 0 in minutes_model).
_ROLE_CHANGE_ROUTING = {
    "lost_starting_spot": ("manager_tendency", "negative"),
    "frozen_out": ("manager_tendency", "negative"),
    "won_starting_spot": ("manager_tendency", "positive"),
    "likely_january_exit": ("transfer_likelihood", None),
    "new_position": ("manager_tendency", "note"),
}


def ingest_role_change(con, wb, ingested_date) -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, "RoleChange"), start=2):
        name = _s(row.get("player") or row.get("subject"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        change_raw = (_s(row.get("change")) or "").lower()
        # tolerate a `change` cell that carries extra prose ("lost_starting_spot (risk of)")
        change = next((k for k in _ROLE_CHANGE_ROUTING if k in change_raw), None)
        routing = _ROLE_CHANGE_ROUTING.get(change) if change else None
        if not uid or not common or routing is None:
            skip += 1
            continue
        claim_type, valence = routing
        payload = {"change": change, "cause": _s(row.get("cause")),
                   "effective_from": _s(row.get("effective_from")), "notes": _s(row.get("notes"))}
        if claim_type == "manager_tendency":
            payload["valence"] = valence
        else:  # transfer_likelihood -- deliberately NOT status "Complete" (that path is
               # skipped by minutes_model as a settled move, not an ongoing minutes risk)
            payload["status"] = "Expected"
        ins += _emit(con, ingested_date, player_uid=uid, claim_type=claim_type,
                     claim_value=payload, common=common, tab="RoleChange", i=i,
                     numeric=1.0 if claim_type == "transfer_likelihood" else None)
    return {"inserted": ins, "skipped": skip}


def ingest_set_pieces(con, wb, ingested_date, tab="SetPieces") -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, tab), start=2):
        common = _common(con, row)
        club = _s(row.get("club"))
        duty = _s(row.get("duty"))
        if not common or not duty:
            skip += 1
            continue
        for col, order in (("primary_taker", "primary"), ("secondary_taker", "secondary")):
            taker = _s(row.get(col) or (row.get("primary") if order == "primary" else row.get("secondary")))
            if not taker or taker in ("-", "n/a", "unclear", "none"):
                continue
            uid = iw._resolve_player(con, taker)
            if not uid:
                skip += 1
                continue
            payload = {"club": club, "duty": duty, "order": order,
                       "deputy_if_primary_absent": _s(row.get("deputy_if_primary_absent")),
                       "notes": _s(row.get("notes"))}
            ins += _emit(con, ingested_date, player_uid=uid, claim_type="set_piece_order_override",
                         claim_value=payload, common=common, tab=tab, i=i)
    return {"inserted": ins, "skipped": skip}


def ingest_price_watch(con, wb, ingested_date, tab="PriceWatch") -> dict:
    ins = skip = 0
    for i, row in enumerate(_rows(wb, tab), start=2):
        name = _s(row.get("player") or row.get("subject"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        if not uid or not common:
            skip += 1
            continue
        payload = {"direction": _s(row.get("direction")),
                   "note": _s(row.get("note") or row.get("claim_value")), "notes": _s(row.get("notes"))}
        numeric = _num(row.get("claim_value_numeric_gbp_m"))
        ins += _emit(con, ingested_date, player_uid=uid, claim_type="fpl_price_note",
                     claim_value=payload, common=common, tab=tab, i=i, numeric=numeric)
    return {"inserted": ins, "skipped": skip}


def ingest_transfers_v1(con, wb, ingested_date) -> dict:
    """v1-only tab. A completed move -- minutes_model treats status='Complete' as a settled
    fact, not an ongoing minutes risk, and skips it; the value here is provenance + the
    chip-timing evidence-freshness check."""
    ins = skip = 0
    for i, row in enumerate(_rows(wb, "Transfers"), start=2):
        name = _s(row.get("subject") or row.get("player") or row.get("subject_entity"))
        common = _common(con, row)
        uid = iw._resolve_player(con, name) if name else None
        if not uid or not common:
            skip += 1
            continue
        payload = {"status": "Complete", "old_club": _s(row.get("club_from")), "new_club": _s(row.get("club_to")),
                   "description": _s(row.get("claim_value")), "notes": _s(row.get("notes"))}
        ins += _emit(con, ingested_date, player_uid=uid, claim_type="transfer_likelihood",
                     claim_value=payload, common=common, tab="Transfers", i=i, numeric=1.0)
    return {"inserted": ins, "skipped": skip}


def ingest_all(con: duckdb.DuckDBPyConnection, xlsx_path: str, source_tier_params_version: int) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    n_sources = register_new_sources(con, wb, source_tier_params_version)
    ingested_date = datetime.now(timezone.utc)
    present = [t for t in _KNOWN_TABS if t in wb.sheetnames]

    out: dict = {"new_sources_registered": n_sources, "tabs_present": present}
    out["injuries"] = ingest_injuries(con, wb, ingested_date)
    out["predicted_xi"] = ingest_predicted_xi(con, wb, ingested_date)
    out["rotation"] = ingest_rotation(con, wb, ingested_date)
    out["role_change"] = ingest_role_change(con, wb, ingested_date)
    out["set_pieces"] = ingest_set_pieces(con, wb, ingested_date, tab="SetPieces" if "SetPieces" in present else "SetPieceTakers")
    out["price_watch"] = ingest_price_watch(con, wb, ingested_date, tab="PriceWatch" if "PriceWatch" in present else "PriceNotes")
    if "Transfers" in present:
        out["transfers_v1"] = ingest_transfers_v1(con, wb, ingested_date)
    return out
