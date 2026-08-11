"""Ingests the evidence workbook into evidence_claims, enforcing the deprecation
allowlist (M0 + M8's tab 31-36 addition) so deprecated tabs are structurally absent
rather than silently skipped by convention.
"""

import json
import re
import uuid
from datetime import datetime, timezone

import duckdb
import openpyxl

# ============================================================
# Tab classification -- this literally IS the enforced allowlist. ingest_all() refuses
# to process any tab not listed here, and 'excluded_deprecated' tabs are never read for
# claims under any circumstance.
# ============================================================

EXCLUDED_DEPRECATED = {
    "1_Player Database": "M0: deprecated raw player-database tab",
    "2_Team Database": "M0: deprecated raw team-database tab",
    "3_Fixture Database": "M0: deprecated raw fixture-database tab",
    "21_Master Player Rating Engine": "M0: deprecated derived rating engine (v1)",
    "21b_Rating Engine Methodology": "M0: methodology tab for deprecated v1 rating engine",
    "28_Master Rating Engine v2": "M0: deprecated derived rating engine (v2)",
    "28b_Rating Engine v2 Method": "M0: methodology tab for deprecated v2 rating engine",
    "29_Position Rankings": "M0: deprecated derived output",
    "30_Fixture Swing GW1-10": "M0: deprecated derived output",
    # M8 addition (kickoff notes item 4): transitively built on 28_Master Rating Engine v2
    "31_Captaincy-Transfer Plan": "M8: downstream of deprecated 28_Master Rating Engine v2",
    "32_Chip Strategy Plan": "M8: downstream of deprecated 28_Master Rating Engine v2",
    "33_10 Squad Variants": "M8: downstream of deprecated 28_Master Rating Engine v2",
    "34_Sensitivity Analysis": "M8: downstream of deprecated 28_Master Rating Engine v2",
    "35_Risk Report": "M8: downstream of deprecated 28_Master Rating Engine v2",
    "36_Top-N Final Lists": "M8: downstream of deprecated 28_Master Rating Engine v2",
}

AUDIT_METADATA = {
    "9_Missing Info Tracker", "10_Data Quality Report", "11_QC Notes (Dedup Log)",
    "12_Questions for Review", "19_Data Quality Issues", "20_Audit Log",
    "22_Audit - Sheet by Sheet", "23_Global Findings", "24_Scores With No Evidence",
}

DOCUMENTATION = {"25_Part 1 Summary & Roadmap", "37_Weekly Update Workflow"}

REFERENCE_ONLY = {
    "7_Source Database": "assorted single-fact log, not structurally tied to a clean player/team/fixture subject column",
    "8_Source Reliability Index": "citation counts, not a reliability score -- feeds sources.citation_count only",
    "13_Rule Changes Database": "feeds M3/M8 rule-verification gates, not player evidence",
    "13b_FPL Rule Database": "feeds M3/M8 rule-verification gates, not player evidence",
    "26_Club Name Map": "team-alias reference, consumed by reconcile.py",
    "27_Rating Weights (Editable)": "feeds the deprecated 28_Master Rating Engine v2; not itself named on the deprecation list, but practically inert once its only consumer is excluded",
}

INGEST_CLAIMS = {
    "4_Injury Database", "5_Transfer Database", "14_Community Evidence",
    "15_Analyst Debate Database", "16_YouTube Evidence Database", "18_Predicted XI Database",
}

INGEST_MANUAL_DECOMPOSITION = {"6_Manager Database", "17_Pre-season Match Reports"}

ALL_CLASSIFIED = (
    set(EXCLUDED_DEPRECATED) | AUDIT_METADATA | DOCUMENTATION | set(REFERENCE_ONLY)
    | INGEST_CLAIMS | INGEST_MANUAL_DECOMPOSITION
)


def seed_allowlist(con: duckdb.DuckDBPyConnection, wb: openpyxl.Workbook) -> None:
    unclassified = set(wb.sheetnames) - ALL_CLASSIFIED
    if unclassified:
        raise ValueError(f"workbook has tabs with no allowlist classification: {unclassified}")
    rows = []
    for tab, reason in EXCLUDED_DEPRECATED.items():
        rows.append((tab, "excluded_deprecated", reason))
    for tab in AUDIT_METADATA:
        rows.append((tab, "audit_metadata", "retained as metadata about evidence quality, not ingested as claims (M0)"))
    for tab in DOCUMENTATION:
        rows.append((tab, "documentation", "narrative/process content, not data"))
    for tab, reason in REFERENCE_ONLY.items():
        rows.append((tab, "reference_only", reason))
    for tab in INGEST_CLAIMS:
        rows.append((tab, "ingest_claims", "structurally atomic rows, mapped directly to evidence_claims"))
    for tab in INGEST_MANUAL_DECOMPOSITION:
        rows.append((tab, "ingest_manual_decomposition", "compound free-text cells; staged for human curation per M1b"))
    for tab_name, status, reason in rows:
        con.execute(
            "INSERT INTO workbook_tab_allowlist (tab_name, status, reason) VALUES (?, ?, ?) "
            "ON CONFLICT (tab_name) DO NOTHING",
            [tab_name, status, reason],
        )


# ============================================================
# sources
# ============================================================

_OFFICIAL_KEYWORDS = ["premier league", "pl official", "fpl official", "clubelo"]
_JOURNALIST_KEYWORDS = [
    "bbc", "guardian", "yahoo", "sky sports", "the athletic", "daily mail", "mirror",
    "telegraph", "espn", "sports illustrated", "independent", "standard",
]
_SPECIALIST_KEYWORDS = [
    "fantasy football fix", "ffs", "fantasy football scout", "rri", "scout", "analyst",
]

_SOURCE_TYPE_OVERRIDES = {
    "PL": "official",
    "Premier League": "official",
    "Premier League Scout (Official)": "official",
    "Premier League official": "official",
    "RRI": "specialist",
    "FFS": "specialist",
}


def classify_source_type(source_name: str) -> str:
    if source_name in _SOURCE_TYPE_OVERRIDES:
        return _SOURCE_TYPE_OVERRIDES[source_name]
    low = source_name.lower()
    if any(k in low for k in _OFFICIAL_KEYWORDS):
        return "official"
    if any(k in low for k in _JOURNALIST_KEYWORDS):
        return "journalist"
    if any(k in low for k in _SPECIALIST_KEYWORDS):
        return "specialist"
    return "community"  # conservative default for anything unrecognized


def _log_scaled(citation_count: float, max_citation_count: float) -> float:
    import math

    if max_citation_count <= 0:
        return 0.0
    return math.log(1 + citation_count) / math.log(1 + max_citation_count)


_CONFIDENCE_BUCKET_RE = re.compile(r"^\d{1,2}/10$")


def build_sources(con: duckdb.DuckDBPyConnection, wb: openpyxl.Workbook, params_version: int) -> int:
    citation_counts: dict[str, int] = {}
    ws = wb["8_Source Reliability Index"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, count = row[0], row[1]
        if not name or not isinstance(count, (int, float)):
            continue
        name = str(name).strip()
        # Real, verified data-quality issue in the source workbook: some rows in this tab
        # are confidence-score buckets ("10/10" with 508 citations, "8/10" with 16, ...),
        # not source names -- they got mixed into the same two-column list with no
        # distinguishing column. Left in, "10/10" would dominate max_citation_count and
        # silently shrink every real source's log-scaled reliability score.
        if _CONFIDENCE_BUCKET_RE.match(name):
            continue
        citation_counts[name] = int(count)
    max_citations = max(citation_counts.values(), default=1)

    source_names: set[str] = set(citation_counts)
    tab_columns = {
        "4_Injury Database": "Source",
        "5_Transfer Database": "Source",
        "14_Community Evidence": "Creator/Source",
        "16_YouTube Evidence Database": "Creator",
        "18_Predicted XI Database": "Source",
    }
    for tab, col in tab_columns.items():
        ws = wb[tab]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if col not in header:
            continue
        idx = header.index(col)
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[idx] if idx < len(row) else None
            if val and isinstance(val, str) and not val.strip().startswith("⚠"):
                source_names.add(val.strip())

    ws = wb["15_Analyst Debate Database"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col_name in ("Creator/Source A", "Creator/Source B"):
        if col_name not in header:
            continue
        idx = header.index(col_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[idx] if idx < len(row) else None
            if val and isinstance(val, str):
                source_names.add(val.strip())

    system_derived = {"system-derived"}
    n = 0
    for name in sorted(source_names | system_derived):
        if name == "system-derived":
            source_type = "system-derived"
            citation_count = None
            base_score = None
        else:
            source_type = classify_source_type(name)
            citation_count = citation_counts.get(name, 1)
            _val_num, tier_weight_text = None, None
            row = con.execute(
                "SELECT value_numeric FROM param_versions WHERE param_family = 'source_tier_weights' "
                "AND param_version = ? AND param_key = 'tier_weight' AND dimensions = ?",
                [params_version, json.dumps({"source_type": source_type}, sort_keys=True, separators=(",", ":"))],
            ).fetchone()
            tier_weight = row[0] if row else 0.4
            base_score = tier_weight * _log_scaled(citation_count, max_citations)
        source_id = "src_" + name.lower().replace(" ", "_").replace("/", "_")
        con.execute(
            "INSERT INTO sources (source_id, source_name, source_type, base_reliability_score, "
            "citation_count, source_notes, last_reviewed_date) VALUES (?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT (source_name) DO NOTHING",
            [source_id, name, source_type, base_score, citation_count, None, None],
        )
        n += 1
    return n


def _source_id_for(con: duckdb.DuckDBPyConnection, name: str | None) -> str | None:
    if not name:
        return None
    row = con.execute("SELECT source_id FROM sources WHERE source_name = ?", [str(name).strip()]).fetchone()
    return row[0] if row else None


def _confidence_0_1(raw) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(raw) / 10.0
    except (TypeError, ValueError):
        return None


def _parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, "year"):  # date
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_player(con: duckdb.DuckDBPyConnection, name: str | None, season: str = "2026-2027") -> str | None:
    """Matches on normalized name, not the literal string -- the workbook spells names
    differently (accents, dropped middle names) than FPL-Core-Insights does. Falls back to
    a word-subset match for compound-surname cases (workbook's "Manuel Ugarte" vs the
    registered "Manuel Ugarte Ribeiro"), but only when it resolves to exactly one distinct
    player, so a short/ambiguous name never silently guesses the wrong one.
    """
    from . import entity_resolution as er

    if not name:
        return None
    norm = er.normalize_name(name)
    if not norm:
        return None

    row = con.execute(
        "SELECT player_uid FROM player_alias WHERE normalized_alias_name = ? AND season = ?",
        [norm, season],
    ).fetchone()
    if row:
        return row[0]
    row = con.execute(
        "SELECT DISTINCT player_uid FROM player_alias WHERE normalized_alias_name = ?", [norm]
    ).fetchone()
    if row:
        return row[0]

    words = set(norm.split())
    if len(words) >= 2:
        candidates = con.execute(
            "SELECT DISTINCT player_uid, normalized_alias_name FROM player_alias"
        ).fetchall()
        matches = {uid for uid, cand_norm in candidates if words.issubset(set(cand_norm.split()))}
        if len(matches) == 1:
            return matches.pop()
    return None


def _insert_claim(con, *, subject_entity_type, subject_entity_id, claim_type, claim_value,
                   claim_value_numeric, information_type, source_id, source_reliability_score,
                   confidence, observed_date, ingested_date, tab_origin, row_origin, raw_text=None):
    if subject_entity_id is None or source_id is None:
        return False
    con.execute(
        "INSERT INTO evidence_claims (claim_id, subject_entity_type, subject_entity_id, claim_type, "
        "claim_value, claim_value_numeric, information_type, source_id, source_reliability_score, "
        "confidence, observed_date, ingested_date, superseded_by, tab_origin, row_origin, raw_text) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?)",
        [
            str(uuid.uuid4()), subject_entity_type, subject_entity_id, claim_type,
            json.dumps(claim_value) if claim_value is not None else None, claim_value_numeric,
            information_type, source_id, source_reliability_score, confidence, observed_date,
            ingested_date, tab_origin, row_origin, raw_text,
        ],
    )
    return True


def _reliability_for(con, source_id):
    row = con.execute("SELECT base_reliability_score FROM sources WHERE source_id = ?", [source_id]).fetchone()
    return row[0] if row and row[0] is not None else 0.0


def ingest_injury_database(con, wb, ingested_date) -> dict:
    ws = wb["4_Injury Database"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        player, club, issue, date_reported, expected_return, gw1_status, info_type, source, source_date, conf, relevance = row[:11]
        player_uid = _resolve_player(con, player)
        source_id = _source_id_for(con, source)
        if not player_uid or not source_id:
            skipped += 1
            continue
        ok = _insert_claim(
            con, subject_entity_type="player", subject_entity_id=player_uid,
            claim_type="injury_status",
            claim_value={"category": gw1_status, "issue": issue, "expected_return": expected_return},
            claim_value_numeric=None, information_type=info_type, source_id=source_id,
            source_reliability_score=_reliability_for(con, source_id), confidence=_confidence_0_1(conf),
            observed_date=_parse_date(source_date) or _parse_date(date_reported), ingested_date=ingested_date,
            tab_origin="4_Injury Database", row_origin=i,
        )
        inserted += 1 if ok else 0
        skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def ingest_transfer_database(con, wb, ingested_date) -> dict:
    ws = wb["5_Transfer Database"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        player, old_club, new_club, fee_type, fee, tdate, status, info_type, source, source_date, conf, relevance = row[:12]
        player_uid = _resolve_player(con, player)
        source_id = _source_id_for(con, source)
        if not player_uid or not source_id:
            skipped += 1
            continue
        ok = _insert_claim(
            con, subject_entity_type="player", subject_entity_id=player_uid,
            claim_type="transfer_likelihood",
            claim_value={"status": status, "old_club": old_club, "new_club": new_club, "fee_type": fee_type},
            claim_value_numeric=1.0 if status == "Complete" else None,
            information_type=info_type, source_id=source_id,
            source_reliability_score=_reliability_for(con, source_id), confidence=_confidence_0_1(conf),
            observed_date=_parse_date(source_date) or _parse_date(tdate), ingested_date=ingested_date,
            tab_origin="5_Transfer Database", row_origin=i,
        )
        inserted += 1 if ok else 0
        skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def ingest_predicted_xi(con, wb, ingested_date) -> dict:
    ws = wb["18_Predicted XI Database"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (label, club, player, position, price, predicted_starter, start_conf, exp_position,
         exp_minutes, rotation_risk, competition_for_pos, backup, preseason_status, system_fit,
         reasoning, conf_level, source, source_date, cross_check) = row[:19]
        player_uid = _resolve_player(con, player)
        source_id = _source_id_for(con, source)
        if not player_uid or not source_id:
            skipped += 1
            continue
        try:
            numeric = float(start_conf) / 100.0 if start_conf not in (None, "") else None
        except (TypeError, ValueError):
            numeric = None
        ok = _insert_claim(
            con, subject_entity_type="player", subject_entity_id=player_uid,
            claim_type="predicted_xi",
            claim_value={
                "label": label, "predicted_starter": predicted_starter, "rotation_risk": rotation_risk,
                "expected_minutes": exp_minutes, "reasoning": reasoning,
            },
            claim_value_numeric=numeric,
            information_type="OPINION",  # tab has no FACT/OPINION column; a predicted lineup is inherently analytical judgment
            source_id=source_id, source_reliability_score=_reliability_for(con, source_id), confidence=None,
            observed_date=_parse_date(source_date), ingested_date=ingested_date,
            tab_origin="18_Predicted XI Database", row_origin=i,
        )
        inserted += 1 if ok else 0
        skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def ingest_community_evidence(con, wb, ingested_date) -> dict:
    ws = wb["14_Community Evidence"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        creator, subject, claim, category, ev_type, supporting, conf, reliability_note, edate, notes = row[:10]
        player_uid = _resolve_player(con, subject)
        source_id = _source_id_for(con, creator)
        if not player_uid or not source_id:
            skipped += 1  # e.g. "Squad strategy" / "Draft squad" -- not a resolvable player/team subject
            continue
        ok = _insert_claim(
            con, subject_entity_type="player", subject_entity_id=player_uid,
            claim_type="community_sentiment",
            claim_value={"claim": claim, "category": category, "notes": notes}, claim_value_numeric=None,
            information_type="FACT" if isinstance(ev_type, str) and ev_type.upper().startswith("FACT") else "OPINION",
            source_id=source_id, source_reliability_score=_reliability_for(con, source_id),
            confidence=_confidence_0_1(conf), observed_date=_parse_date(edate), ingested_date=ingested_date,
            tab_origin="14_Community Evidence", row_origin=i,
        )
        inserted += 1 if ok else 0
        skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def ingest_analyst_debate(con, wb, ingested_date) -> dict:
    """M1b: each row splits into two ordinary evidence_claims rows (Opinion A, Opinion B)."""
    ws = wb["15_Analyst Debate Database"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (player_field, src_a, opinion_a, src_b, opinion_b, reason_a, reason_b, supporting,
         contradicting, strength, notes) = row[:11]
        names = [n.strip() for n in str(player_field).split(" vs ")] if player_field else []
        single = _resolve_player(con, player_field)
        if single:
            name_a = name_b = player_field
        elif len(names) == 2:
            name_a, name_b = names
        else:
            skipped += 2
            continue
        for side_name, src, opinion, reason in (
            (name_a, src_a, opinion_a, reason_a), (name_b, src_b, opinion_b, reason_b)
        ):
            player_uid = _resolve_player(con, side_name)
            source_id = _source_id_for(con, src)
            if not player_uid or not source_id:
                skipped += 1
                continue
            ok = _insert_claim(
                con, subject_entity_type="player", subject_entity_id=player_uid,
                claim_type="analyst_debate",
                claim_value={"opinion": opinion, "reason": reason, "strength": strength},
                claim_value_numeric=None, information_type="OPINION",
                source_id=source_id, source_reliability_score=_reliability_for(con, source_id), confidence=None,
                observed_date=None, ingested_date=ingested_date,
                tab_origin="15_Analyst Debate Database", row_origin=i,
            )
            inserted += 1 if ok else 0
            skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def ingest_youtube_evidence(con, wb, ingested_date) -> dict:
    ws = wb["16_YouTube Evidence Database"]
    inserted, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        creator, video, edate, player, club, position, claim, ev_type, supporting, conf, reliability = row[:11]
        if not creator or (isinstance(creator, str) and creator.strip().startswith("⚠")):
            continue  # placeholder "no claims in this batch" rows
        player_uid = _resolve_player(con, player)
        source_id = _source_id_for(con, creator)
        if not player_uid or not source_id:
            skipped += 1
            continue
        ok = _insert_claim(
            con, subject_entity_type="player", subject_entity_id=player_uid,
            claim_type="youtube_evidence",
            claim_value={"claim": claim, "supporting_data": supporting}, claim_value_numeric=None,
            information_type="FACT" if isinstance(ev_type, str) and ev_type.upper().startswith("FACT") else "OPINION",
            source_id=source_id, source_reliability_score=_reliability_for(con, source_id),
            confidence=_confidence_0_1(conf), observed_date=_parse_date(edate), ingested_date=ingested_date,
            tab_origin="16_YouTube Evidence Database", row_origin=i,
        )
        inserted += 1 if ok else 0
        skipped += 0 if ok else 1
    return {"inserted": inserted, "skipped": skipped}


def stage_manager_database(con, wb, ingested_date) -> int:
    ws = wb["6_Manager Database"]
    n = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        club, note, info_type, source, source_date, conf, relevance = row[:7]
        if not note:
            continue
        source_id = _source_id_for(con, source)
        con.execute(
            "INSERT INTO claims_pending_manual_decomposition (subject_hint, raw_text, information_type, "
            "source_id, source_date, confidence_raw, fpl_relevance, tab_origin, row_origin, ingested_date) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [club, note, info_type, source_id, _parse_date(source_date), conf, relevance,
             "6_Manager Database", i, ingested_date],
        )
        n += 1
    return n


def stage_preseason_match_reports(con, wb, ingested_date) -> int:
    ws = wb["17_Pre-season Match Reports"]
    n = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fixture, score, scorers, assists, notes, ev_type, source, conf = row[:8]
        if not notes:
            continue
        source_id = _source_id_for(con, source)
        con.execute(
            "INSERT INTO claims_pending_manual_decomposition (subject_hint, raw_text, information_type, "
            "source_id, source_date, confidence_raw, fpl_relevance, tab_origin, row_origin, ingested_date) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [fixture, notes, ev_type, source_id, None, conf, None,
             "17_Pre-season Match Reports", i, ingested_date],
        )
        n += 1
    return n


def ingest_all(con: duckdb.DuckDBPyConnection, xlsx_path: str, source_tier_params_version: int) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    seed_allowlist(con, wb)
    n_sources = build_sources(con, wb, source_tier_params_version)
    ingested_date = datetime.now(timezone.utc)

    results = {
        "sources": n_sources,
        "injury": ingest_injury_database(con, wb, ingested_date),
        "transfer": ingest_transfer_database(con, wb, ingested_date),
        "predicted_xi": ingest_predicted_xi(con, wb, ingested_date),
        "community": ingest_community_evidence(con, wb, ingested_date),
        "analyst_debate": ingest_analyst_debate(con, wb, ingested_date),
        "youtube": ingest_youtube_evidence(con, wb, ingested_date),
        "manager_db_staged": stage_manager_database(con, wb, ingested_date),
        "preseason_reports_staged": stage_preseason_match_reports(con, wb, ingested_date),
    }
    return results
